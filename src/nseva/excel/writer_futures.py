"""FuData Excel rendering (Implementation Plan §6.8 & §9)."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from nseva.gold.futures_summary import build_futures_summary
from nseva.services.expiry_service import windows_for
from nseva.util.paths import data_root_from_config

try:
    from nseva.config import load_config
except Exception:  # pragma: no cover - allow early imports
    load_config = None  # type: ignore[assignment]


BLOCK_ROWS = 75
BLOCK_COLS = 16  # A:P in the template


def render_futures_workbook(
    symbol: str,
    expiries: Sequence[date],
    *,
    template_path: Path,
    output_path: Path,
    storage_root: Path | str | None = None,
) -> None:
    """Render/update the FuData workbook for `symbol` using the provided template."""

    cfg = load_config() if load_config else None
    root = data_root_from_config(storage_root or (cfg.runtime.storage_root if cfg else "./data"))

    template_wb = load_workbook(template_path)
    template_ws = template_wb.active
    sheet_name = template_ws.title

    template_cells = _capture_block(template_ws, BLOCK_ROWS, BLOCK_COLS)
    template_merges = _capture_merges(template_ws, BLOCK_ROWS, BLOCK_COLS)
    label_positions = _locate_labels(template_ws)
    table_headers = _extract_table_headers(template_ws, header_row=5, start_col=1, max_cols=BLOCK_COLS)

    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    spacing_cols = getattr(cfg.futures, "table_spacing_rows", 2) if cfg else 2
    start_row = 1
    start_col = 1

    for idx, expiry in enumerate(sorted(expiries)):
        dest_col = start_col + idx * (BLOCK_COLS + spacing_cols)
        if idx > 0:
            _apply_template_block(ws, template_cells, dest_row=start_row, dest_col=dest_col)
            _apply_merges(ws, template_merges, dest_row=start_row, dest_col=dest_col)
        _render_block(
            ws,
            symbol,
            expiry,
            start_row,
            dest_col,
            root,
            label_positions,
            table_headers,
            cfg,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _render_block(
    ws: Worksheet,
    symbol: str,
    expiry: date,
    row: int,
    col: int,
    root: Path,
    labels: Mapping[str, tuple[int, int]],
    table_headers: list[tuple[int, int, str]],
    cfg: Any,
) -> None:
    """Render a single monthly block."""

    primary_start, overlap_start, _ = windows_for(symbol, expiry, storage_root=root)
    summary_df = build_futures_summary(symbol, expiry, storage_root=root)
    gold_df = _load_gold_for_window(symbol, overlap_start, expiry, root=root)

    _set_if_exists(ws, labels, "Symbol =", symbol, col_offset=1, row=row, col=col)
    lot_size = int(gold_df["lot_size_shares"].iloc[0]) if not gold_df.empty else None
    _set_if_exists(ws, labels, "Lot Size =", lot_size, col_offset=1, row=row, col=col)
    _set_if_exists(ws, labels, "Expiry =", expiry, col_offset=1, row=row, col=col)
    _set_if_exists(ws, labels, "Start dt. =", primary_start, col_offset=1, row=row, col=col)
    _set_if_exists(ws, labels, "Overlap Start dt. =", overlap_start, col_offset=1, row=row, col=col)

    if not summary_df.empty:
        _set_if_exists(
            ws,
            labels,
            "Max. Contracts =",
            _safe_excel_number(summary_df["max_permitted_contracts"].iloc[0]),
            col_offset=1,
            row=row,
            col=col,
        )
        _set_if_exists(
            ws,
            labels,
            "90% of Max. OI Contracts =",
            _safe_excel_number(summary_df["threshold_90pct"].iloc[0]),
            col_offset=1,
            row=row,
            col=col,
        )
        _set_if_exists(
            ws,
            labels,
            "Max. OI Contracts month =",
            _safe_excel_number(summary_df["max_oi_contracts"].iloc[0]),
            col_offset=1,
            row=row,
            col=col,
        )

    if table_headers:
        _write_daily_table(
            ws,
            gold_df,
            table_headers,
            start_row=row + table_headers[0][0],
            start_col=col,
            cfg=cfg,
            symbol=symbol,
        )


def _capture_block(ws: Worksheet, rows: int, cols: int) -> list[dict[str, object]]:
    """Capture values/styles/formulas for a template block."""

    captured: list[dict[str, object]] = []
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None and not cell.has_style:
                continue
            captured.append(
                {
                    "row": r,
                    "col": c,
                    "value": cell.value,
                    "style": copy(cell._style),
                }
            )
    return captured


def _capture_merges(ws: Worksheet, rows: int, cols: int) -> list[tuple[int, int, int, int]]:
    """Capture merged cell ranges within the template block."""

    merges: list[tuple[int, int, int, int]] = []
    for merge in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        if min_row <= rows and min_col <= cols:
            merges.append((min_row, min_col, max_row, max_col))
    return merges


def _apply_template_block(
    ws: Worksheet, template_cells: list[dict[str, object]], *, dest_row: int, dest_col: int
) -> None:
    """Apply captured template block to a new location translating formulas."""

    for entry in template_cells:
        src_row = entry["row"]
        src_col = entry["col"]
        value = entry["value"]
        style = entry["style"]
        target = ws.cell(row=dest_row + src_row - 1, column=dest_col + src_col - 1)

        if isinstance(value, str) and value.startswith("="):
            origin_coord = f"{get_column_letter(src_col)}{src_row}"
            target_coord = f"{get_column_letter(dest_col + src_col - 1)}{dest_row + src_row - 1}"
            target.value = Translator(value, origin=origin_coord).translate_formula(target_coord)
        else:
            target.value = value

        target._style = copy(style)


def _apply_merges(ws: Worksheet, merges: list[tuple[int, int, int, int]], *, dest_row: int, dest_col: int) -> None:
    """Apply captured merged ranges to the destination block."""

    for min_row, min_col, max_row, max_col in merges:
        ws.merge_cells(
            start_row=dest_row + min_row - 1,
            start_column=dest_col + min_col - 1,
            end_row=dest_row + max_row - 1,
            end_column=dest_col + max_col - 1,
        )


def _locate_labels(ws: Worksheet) -> dict[str, tuple[int, int]]:
    """Find key label positions relative to the block origin (1,1)."""

    targets = {
        "Symbol =",
        "Lot Size =",
        "Expiry =",
        "Start dt. =",
        "Overlap Start dt. =",
        "Max. Contracts =",
        "90% of Max. OI Contracts =",
        "Max. OI Contracts month =",
    }
    found: dict[str, tuple[int, int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                stripped = cell.value.strip()
                if stripped in targets:
                    found[stripped] = (cell.row, cell.column)
    return found


def _extract_table_headers(
    ws: Worksheet, *, header_row: int, start_col: int, max_cols: int
) -> list[tuple[int, int, str]]:
    """Extract header texts and positions from the template relative to block origin."""

    headers: list[tuple[int, int, str]] = []
    for col in range(start_col, start_col + max_cols):
        cell = ws.cell(row=header_row, column=col)
        if isinstance(cell.value, str) and cell.value.strip():
            headers.append((header_row - 1, col - 1, cell.value.strip()))
    return headers


def _write_daily_table(
    ws: Worksheet,
    gold_df: pd.DataFrame,
    headers: list[tuple[int, int, str]],
    *,
    start_row: int,
    start_col: int,
    cfg: Any,
    symbol: str,
) -> None:
    """Populate the W3 daily table using template header mapping."""

    if gold_df.empty:
        return

    gold_df = gold_df.sort_values("trade_date")
    rounding = cfg.display_rounding.digits if cfg else None

    header_map = {
        "Instrument": lambda rec: "FUTSTK",
        "Symbol": lambda rec: symbol,
        "Date": lambda rec: rec["trade_date"],
        "Expiry": lambda rec: rec["expiry_date"],
        "Open": lambda rec: rec["open"],
        "High": lambda rec: rec["high"],
        "Low": lambda rec: rec["low"],
        "Close": lambda rec: rec["close"],
        "LTP": lambda rec: rec.get("settle_price"),
        "Settle Price": lambda rec: rec.get("settle_price"),
        "No. of contracts": lambda rec: rec.get("contracts"),
        "Turnover in lacs": lambda rec: rec.get("value_lakhs"),
        "Open Int": lambda rec: rec.get("oi_contracts"),
        "OI Contracts": lambda rec: rec.get("oi_contracts"),
        "Change in OI": lambda rec: rec.get("change_in_oi_contracts"),
        " Change in OI Contracts": lambda rec: rec.get("change_in_oi_contracts"),
    }

    for row_idx, (_, rec) in enumerate(gold_df.iterrows(), start=1):
        for _, header_col, header_text in headers:
            if header_text not in header_map:
                continue
            raw_value = header_map[header_text](rec)
            value = _apply_rounding(header_text, raw_value, rounding)
            ws.cell(row=start_row + row_idx, column=start_col + header_col, value=value)


def _load_gold_for_window(symbol: str, start: date, end: date, *, root: Path) -> pd.DataFrame:
    """Load gold futures_day rows for a window."""

    gold_root = root / "gold" / "futures_day"
    frames: list[pd.DataFrame] = []
    for partition in gold_root.glob("date=*"):
        date_str = partition.name.split("=", 1)[1]
        trade_date = date.fromisoformat(date_str)
        if trade_date < start or trade_date > end:
            continue
        path = partition / f"{symbol}.parquet"
        if path.exists():
            df = pd.read_parquet(path)
            frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _set_if_exists(
    ws: Worksheet,
    labels: Mapping[str, tuple[int, int]],
    label: str,
    value: object,
    *,
    col_offset: int,
    row: int,
    col: int,
) -> None:
    """Set value adjacent to a known label if present."""

    if label not in labels:
        return
    rel_row, rel_col = labels[label]
    target_row = row + rel_row - 1
    target_col = col + rel_col - 1 + col_offset
    anchor_row, anchor_col = _anchor_for_cell(ws, target_row, target_col)
    ws.cell(row=anchor_row, column=anchor_col, value=value)


def _apply_rounding(header: str, value: object, rounding: Any) -> object:
    """Apply display rounding per config digits."""

    if rounding is None or not isinstance(value, (int, float)):
        return value

    prices_labels = {"Open", "High", "Low", "Close", "LTP", "Settle Price"}
    contracts_labels = {"No. of contracts", "Open Int", "OI Contracts", "Change in OI", " Change in OI Contracts"}
    quantities_labels = {"Turnover in lacs"}

    if header in prices_labels:
        digits = rounding.prices
    elif header in contracts_labels:
        digits = rounding.contracts
    elif header in quantities_labels:
        digits = rounding.quantities
    else:
        return value

    return round(value, digits)


def _anchor_for_cell(ws: Worksheet, row: int, col: int) -> tuple[int, int]:
    """Return the writable anchor for a potentially merged cell coordinate."""

    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            return merge.min_row, merge.min_col
    return row, col


def _safe_excel_number(value: object) -> object:
    """Openpyxl cannot write pandas.NA/NaT directly."""

    if value is pd.NA:
        return None
    return value


__all__ = ["render_futures_workbook"]
