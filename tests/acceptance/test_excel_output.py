from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from tests.helpers import SYMBOL, run_sample_pipeline


def _coerce_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def _value_right_of(ws, label: str):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                col = cell.column + 1
                while ws.cell(row=cell.row, column=col).value is None and col <= cell.column + 4:
                    col += 1
                return ws.cell(row=cell.row, column=col).value
    raise AssertionError(f"Label {label!r} not found")


def _header_cell(ws, text: str):
    for row in ws.iter_rows(min_row=5, max_row=5):
        for cell in row:
            if cell.value == text:
                return cell
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell
    raise AssertionError(f"Header {text!r} not found")


def test_excel_values_follow_template(monkeypatch, tmp_path) -> None:
    root, expected = run_sample_pipeline(monkeypatch, tmp_path)

    workbook_path = root / "excel" / SYMBOL / "FuData.xlsx"
    assert workbook_path.exists()

    wb = load_workbook(workbook_path)
    ws = wb.active

    assert _coerce_date(_value_right_of(ws, "Start dt. =")) == expected["primary_start"]
    assert _coerce_date(_value_right_of(ws, "Overlap Start dt. =")) == expected["overlap_start"]
    assert _coerce_date(_value_right_of(ws, "Expiry =")) == expected["expiry"]

    assert _value_right_of(ws, "Max. Contracts =") == expected["max_permitted"]
    assert _value_right_of(ws, "90% of Max. OI Contracts =") == expected["threshold_90pct"]
    assert _value_right_of(ws, "Max. OI Contracts month =") == expected["max_oi_contracts"]

    header_date = _header_cell(ws, "Date")
    header_oi = _header_cell(ws, "OI Contracts")
    first_row = header_date.row + 1
    first_trade = _coerce_date(ws.cell(row=first_row, column=header_date.column).value)
    first_oi = ws.cell(row=first_row, column=header_oi.column).value

    assert first_trade == expected["first_trade_date"]
    assert first_oi == expected["oi_by_date"][expected["first_trade_date"].isoformat()]
